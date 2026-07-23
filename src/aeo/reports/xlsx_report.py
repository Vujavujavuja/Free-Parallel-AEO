"""XLSX report: 8-sheet workbook with live Excel formulas for all aggregations
(PRD FR-21, AC-4). The 'Mention Heatmap' sheet is the anchor that the Overview
and Question Aggregate sheets reference via SUM/COUNTIF/INDEX formulas, so the
numbers stay live and consistent when opened in Excel."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from aeo.analysis import AnalysisResult
from aeo.reports.base import get_analysis, sanitize_cell
from aeo.schemas.run import RunRecord

SHEET_HEATMAP = "Mention Heatmap"

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="2F5597")
_BRAND_FILL = PatternFill("solid", fgColor="C6EFCE")
_WRAP = Alignment(wrap_text=True, vertical="top")


def write_xlsx(record: RunRecord, path: Path) -> Path:
    analysis = get_analysis(record)
    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet

    # Heatmap first: other sheets reference its cells.
    _heatmap_sheet(wb, analysis)
    _overview_sheet(wb, record, analysis)
    _question_aggregate_sheet(wb, record, analysis)
    _sources_sheet(wb, analysis)
    _domain_frequency_sheet(wb, analysis)
    _competitor_sov_sheet(wb, analysis)
    _search_queries_sheet(wb, analysis)
    _insights_sheet(wb, analysis)

    # Put Overview first for readers.
    wb.move_sheet("Overview", -wb.sheetnames.index("Overview"))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def _header(ws: Worksheet, labels: list[str]) -> None:
    for col, label in enumerate(labels, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    ws.freeze_panes = "A2"


def _models(analysis: AnalysisResult) -> list[str]:
    return [ma.model_id for ma in analysis.models]


# --- Mention Heatmap (anchor sheet) -----------------------------------------

def _heatmap_sheet(wb: Workbook, analysis: AnalysisResult) -> None:
    ws = wb.create_sheet(SHEET_HEATMAP)
    models = _models(analysis)
    q_indices = analysis.question_indices
    labels = ["Model", *[f"Q{i}" for i in q_indices], "Total"]
    _header(ws, labels)

    n_q = len(q_indices)
    first_q = get_column_letter(2)
    last_q = get_column_letter(1 + n_q) if n_q else first_q
    total_col = 2 + n_q

    for i, ma in enumerate(analysis.models):
        row = 2 + i
        ws.cell(row=row, column=1, value=sanitize_cell(ma.model_id))
        for jpos, qidx in enumerate(q_indices):
            ws.cell(row=row, column=2 + jpos, value=ma.per_question_brand.get(qidx, 0))
        if n_q:
            ws.cell(
                row=row, column=total_col,
                value=f"=SUM({first_q}{row}:{last_q}{row})",
            )

    # Column totals row.
    if models:
        total_row = 2 + len(models)
        ws.cell(row=total_row, column=1, value="Total").font = Font(bold=True)
        for jpos in range(n_q):
            col = get_column_letter(2 + jpos)
            ws.cell(
                row=total_row, column=2 + jpos,
                value=f"=SUM({col}2:{col}{1 + len(models)})",
            )
        if n_q:
            tcol = get_column_letter(total_col)
            ws.cell(
                row=total_row, column=total_col,
                value=f"=SUM({tcol}2:{tcol}{1 + len(models)})",
            )
    _autosize(ws)


# --- Overview ----------------------------------------------------------------

def _overview_sheet(wb: Workbook, record: RunRecord, analysis: AnalysisResult) -> None:
    ws = wb.create_sheet("Overview")
    _header(ws, [
        "Rank", "Model", "Provider", "Used Web Search", "#Searches",
        "Brand Mentions", "Sub-brand Mentions", "In Vendor Table",
        "#Questions Mentioning", "Unique Cite Domains", "Answer Length",
        "#Competitors Mentioned", "Provenance", "Notes",
    ])
    models = _models(analysis)
    n_q = len(analysis.question_indices)
    last_q = get_column_letter(1 + n_q) if n_q else "B"
    hrow = {m: 2 + i for i, m in enumerate(models)}

    for rank, ma in enumerate(analysis.ranked_models(), start=1):
        row = rank + 1
        r = hrow.get(ma.model_id, 2)
        brand_formula = (
            f"=SUM('{SHEET_HEATMAP}'!B{r}:{last_q}{r})" if n_q else 0
        )
        qmention_formula = (
            f'=COUNTIF(\'{SHEET_HEATMAP}\'!B{r}:{last_q}{r},">0")' if n_q else 0
        )
        n_comp = sum(1 for v in ma.competitor_totals.values() if v > 0)
        values = [
            rank,
            sanitize_cell(ma.model_id),
            sanitize_cell(ma.provider),
            "Yes" if ma.web_search_used else "No",
            ma.num_searches,
            brand_formula,
            ma.subbrand_mentions,
            "Yes" if ma.in_vendor_table else "No",
            qmention_formula,
            len(ma.unique_domains),
            ma.answer_length,
            n_comp,
            ma.provenance.value,
            sanitize_cell(ma.error or ""),
        ]
        for col, val in enumerate(values, start=1):
            ws.cell(row=row, column=col, value=val)
    _autosize(ws)


# --- Question Aggregate ------------------------------------------------------

def _question_aggregate_sheet(
    wb: Workbook, record: RunRecord, analysis: AnalysisResult
) -> None:
    ws = wb.create_sheet("Question Aggregate")
    _header(ws, [
        "Q#", "Question", "Total Mentions", "#Models Mentioning",
        "Avg/Model", "Peak Model", "Interpretation",
    ])
    models = _models(analysis)
    m = len(models)
    for jpos, agg in enumerate(analysis.questions):
        row = jpos + 2
        qcol = get_column_letter(2 + jpos)
        rng = f"'{SHEET_HEATMAP}'!{qcol}2:{qcol}{1 + m}" if m else ""
        name_rng = f"'{SHEET_HEATMAP}'!$A$2:$A${1 + m}" if m else ""
        total_cell = f"C{row}"
        if m:
            total_v: object = f"=SUM({rng})"
            models_v: object = f'=COUNTIF({rng},">0")'
            avg_v: object = f"=IF({m}=0,0,{total_cell}/{m})"
            peak_v: object = (
                f'=IFERROR(IF(MAX({rng})=0,"-",'
                f"INDEX({name_rng},MATCH(MAX({rng}),{rng},0))),\"-\")"
            )
        else:
            total_v = models_v = avg_v = 0
            peak_v = "-"
        ws.cell(row=row, column=1, value=agg.index)
        ws.cell(row=row, column=2, value=sanitize_cell(agg.text)).alignment = _WRAP
        ws.cell(row=row, column=3, value=total_v)
        ws.cell(row=row, column=4, value=models_v)
        ws.cell(row=row, column=5, value=avg_v)
        ws.cell(row=row, column=6, value=peak_v)
        ws.cell(row=row, column=7, value=sanitize_cell(agg.interpretation)).alignment = _WRAP
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["G"].width = 40


# --- Sources by Model & Question --------------------------------------------

def _sources_sheet(wb: Workbook, analysis: AnalysisResult) -> None:
    ws = wb.create_sheet("Sources by Model & Q")
    _header(ws, [
        "Model", "Question", "Hyperlinked Cite Domains",
        "Self-reported cited", "Self-reported named", "Example URL",
    ])
    row = 2
    for ma in analysis.models:
        cites_by_q: dict[int, list[tuple[str, str]]] = {}
        for c in ma.citations:
            cites_by_q.setdefault(c.question_index or 0, []).append((c.domain, c.url))
        indices = sorted(set(cites_by_q) | set(ma.self_reported))
        for idx in indices:
            pairs = cites_by_q.get(idx, [])
            sr = ma.self_reported.get(idx, {"cited": [], "named": []})
            domains = ", ".join(sorted({d for d, _ in pairs}))
            ws.cell(row=row, column=1, value=sanitize_cell(ma.model_id))
            ws.cell(row=row, column=2, value=f"Q{idx}" if idx else "Overall")
            ws.cell(row=row, column=3, value=sanitize_cell(domains)).alignment = _WRAP
            c4 = ws.cell(row=row, column=4, value=sanitize_cell(", ".join(sr["cited"])))
            c4.alignment = _WRAP
            c5 = ws.cell(row=row, column=5, value=sanitize_cell(", ".join(sr["named"])))
            c5.alignment = _WRAP
            if pairs:
                link = ws.cell(row=row, column=6, value=pairs[0][1])
                link.hyperlink = pairs[0][1]
                link.font = Font(color="0563C1", underline="single")
            row += 1
    for col, w in {"C": 34, "D": 34, "E": 34, "F": 34}.items():
        ws.column_dimensions[col].width = w


# --- Domain Frequency --------------------------------------------------------

def _domain_frequency_sheet(wb: Workbook, analysis: AnalysisResult) -> None:
    ws = wb.create_sheet("Domain Frequency")
    _header(ws, ["Domain", "#Models Citing", "Models", "Brand-owned?", "Reference?"])
    for i, d in enumerate(analysis.domain_frequency):
        row = i + 2
        ws.cell(row=row, column=1, value=sanitize_cell(d.domain))
        ws.cell(row=row, column=2, value=d.num_models)
        ws.cell(row=row, column=3, value=sanitize_cell(", ".join(d.models))).alignment = _WRAP
        ws.cell(row=row, column=4, value="Yes" if d.brand_owned else "No")
        ws.cell(row=row, column=5, value="Yes" if d.is_reference else "No")
        if d.brand_owned:
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = _BRAND_FILL
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["C"].width = 50


# --- Competitor Share of Voice ----------------------------------------------

def _competitor_sov_sheet(wb: Workbook, analysis: AnalysisResult) -> None:
    ws = wb.create_sheet("Competitor SoV")
    competitors = analysis.competitors
    _header(ws, ["Model", *competitors, "Total"])
    models = _models(analysis)
    n_c = len(competitors)
    first_c = get_column_letter(2)
    last_c = get_column_letter(1 + n_c) if n_c else first_c
    total_col = 2 + n_c

    for i, ma in enumerate(analysis.models):
        row = 2 + i
        ws.cell(row=row, column=1, value=sanitize_cell(ma.model_id))
        for jpos, comp in enumerate(competitors):
            ws.cell(row=row, column=2 + jpos, value=ma.competitor_totals.get(comp, 0))
        if n_c:
            ws.cell(row=row, column=total_col, value=f"=SUM({first_c}{row}:{last_c}{row})")

    if models and n_c:
        total_row = 2 + len(models)
        ws.cell(row=total_row, column=1, value="Total").font = Font(bold=True)
        for jpos in range(n_c):
            col = get_column_letter(2 + jpos)
            ws.cell(row=total_row, column=2 + jpos, value=f"=SUM({col}2:{col}{1 + len(models)})")
        tcol = get_column_letter(total_col)
        ws.cell(row=total_row, column=total_col, value=f"=SUM({tcol}2:{tcol}{1 + len(models)})")
    _autosize(ws)


# --- Search Queries ----------------------------------------------------------

def _search_queries_sheet(wb: Workbook, analysis: AnalysisResult) -> None:
    ws = wb.create_sheet("Search Queries")
    _header(ws, ["Model", "#", "Query"])
    row = 2
    for ma in analysis.models:
        for n, q in enumerate(ma.search_queries, start=1):
            ws.cell(row=row, column=1, value=sanitize_cell(ma.model_id))
            ws.cell(row=row, column=2, value=n)
            ws.cell(row=row, column=3, value=sanitize_cell(q))
            row += 1
    ws.column_dimensions["C"].width = 50


# --- Insights & Quotes -------------------------------------------------------

def _insights_sheet(wb: Workbook, analysis: AnalysisResult) -> None:
    ws = wb.create_sheet("Insights & Quotes")
    ws.cell(row=1, column=1, value="Findings").font = Font(bold=True, size=13)
    row = 2
    for insight in analysis.insights:
        ws.cell(row=row, column=1, value=sanitize_cell(f"• {insight}"))
        row += 1
    row += 1
    ws.cell(row=row, column=1, value="Representative Quotes").font = Font(bold=True, size=13)
    row += 1
    ws.cell(row=row, column=1, value="Model").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Quote").font = Font(bold=True)
    row += 1
    for q in analysis.quotes:
        ws.cell(row=row, column=1, value=sanitize_cell(q.get("model", "")))
        ws.cell(row=row, column=2, value=sanitize_cell(q.get("quote", ""))).alignment = _WRAP
        row += 1
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 80


def _autosize(ws: Worksheet) -> None:
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[letter].width = min(max(length + 2, 10), 40)
