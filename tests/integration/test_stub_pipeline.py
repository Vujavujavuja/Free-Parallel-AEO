"""End-to-end pipeline test on the free stub provider (PRD AC-3..AC-5)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from aeo.constants import Provenance, RunStatus
from aeo.schemas.company import CompanyProfile
from aeo.schemas.run import RunOptions, RunRecord
from aeo.services import run_service
from aeo.storage import RunStore

_ERROR_LITERALS = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!"}
_EXPECTED_SHEETS = {
    "Overview", "Mention Heatmap", "Question Aggregate", "Sources by Model & Q",
    "Domain Frequency", "Competitor SoV", "Search Queries", "Insights & Quotes",
}


async def test_full_stub_run(
    tmp_path: Path, company: CompanyProfile, stub_options: RunOptions
) -> None:
    store = RunStore(tmp_path / "runs")
    record = RunRecord(company=company, options=stub_options)

    result = await run_service.execute_run(
        record, provider_name="stub", store=store,
    )

    # Pipeline completed with the requested question count and a response per model.
    assert result.status == RunStatus.COMPLETED
    assert len(result.questions) == stub_options.question_count
    assert {r.model_id for r in result.responses} == set(stub_options.target_models)
    assert result.total_cost_usd > 0

    # Analysis produced provenance for every model and a heatmap.
    assert result.analysis is not None
    provenances = {m["provenance"] for m in result.analysis["models"]}
    assert provenances <= {p.value for p in Provenance}
    assert result.analysis["heatmap"]

    # All three artifacts exist.
    for fmt in ("xlsx", "csv", "json"):
        assert Path(result.reports[fmt]).is_file()

    # XLSX has all 8 sheets and no pre-stored error literals.
    wb = load_workbook(result.reports["xlsx"])
    assert set(wb.sheetnames) == _EXPECTED_SHEETS
    formula_cells = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str):
                    assert c.value not in _ERROR_LITERALS
                    if c.value.startswith("="):
                        formula_cells += 1
    assert formula_cells > 0  # aggregations use live formulas


async def test_run_is_persisted_and_listable(
    tmp_path: Path, company: CompanyProfile, stub_options: RunOptions
) -> None:
    store = RunStore(tmp_path / "runs")
    record = RunRecord(company=company, options=stub_options)
    await run_service.execute_run(record, provider_name="stub", store=store)

    summaries = run_service.list_runs(store)
    assert len(summaries) == 1
    assert summaries[0].id == record.id

    reloaded = run_service.get_run(record.id, store)
    assert reloaded.status == RunStatus.COMPLETED
